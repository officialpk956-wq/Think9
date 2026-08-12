"""L1 extraction: turn 12 differently-formatted vendor artifacts into LineItem
candidates.

Two families of parser:
  - Tabular (PDF tables, CSV, xlsx): columns are matched by keyword, not
    position, because no two vendors in data/quotes use the same header
    names. One `_map_columns` + `_rows_to_line_items` pair serves all three
    formats.
  - Prose (email, WhatsApp): commercial terms (qty/price/MOQ/lead time) are
    pulled from free text with a shared regex pass; there is no schema to
    match columns against.

Brand and vendor identity come from data/quotes/manifest.json, not from the
document body - that's L0 provenance (who this artifact arrived from),
which a real ingest pipeline attaches at the inbox/channel level rather
than parsing out of the quote text itself.
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Optional

import pdfplumber
from openpyxl import load_workbook

from loom.models import LineItem
from loom.normalize import AttributeSignature, extract_attributes

QUOTES_DIR = Path(__file__).parent.parent / "data" / "quotes"

# Column-name keywords, checked in this priority order. Substring match,
# case-insensitive - "Order Qty", "Qty Requested" and "units" all resolve
# to the same field without hardcoding any vendor's exact header text.
COLUMN_KEYWORDS = [
    ("sku_code", ["sku", "code"]),
    ("description", ["desc", "item", "product"]),
    ("qty", ["qty", "quantity", "units"]),
    ("unit_price", ["rate", "price", "inr"]),
    ("moq", ["moq", "min"]),
    ("lead_time_days", ["lead", "delivery"]),
]

_SKU_CODE_RE = re.compile(r"\b[A-Z]{2,}(?:-[A-Z0-9]+){1,}\b")
_CHAT_TAG_RE = re.compile(r"\[.*?\]:\s*")
_QTY_RE = re.compile(r"(\d[\d,]*)\s*pcs\b", re.I)
_PRICE_RE = re.compile(r"Rs\.?\s*([\d,]+(?:\.\d+)?)\s*(?:/|per\s+)(?:pc|piece|unit)", re.I)
_MOQ_RE = re.compile(r"moq\D{0,40}?(\d[\d,]*)", re.I)
_LEAD_RE = re.compile(r"lead\s*time\D{0,40}?(\d+)\s*days?", re.I)


def _parse_int(raw) -> Optional[int]:
    digits = re.sub(r"[^\d]", "", str(raw)) if raw is not None else ""
    return int(digits) if digits else None


def _parse_number(raw) -> Optional[float]:
    digits = re.sub(r"[^\d.]", "", str(raw)) if raw is not None else ""
    return float(digits) if digits else None


def _confidence(description: str, qty: Optional[int], unit_price: Optional[float]) -> float:
    """Fraction of the fields a usable line item needs that were actually found."""
    return (0.4 if description else 0) + (0.3 if qty is not None else 0) + (0.3 if unit_price is not None else 0)


def _map_columns(header: list) -> dict[str, int]:
    lowered = [str(h or "").lower() for h in header]
    mapping: dict[str, int] = {}
    for field, keywords in COLUMN_KEYWORDS:
        for idx, cell in enumerate(lowered):
            if idx in mapping.values():
                continue
            if any(kw in cell for kw in keywords):
                mapping[field] = idx
                break
    return mapping


def _rows_to_line_items(header, rows, vendor_gstin, brands, source_file, source_format) -> list[LineItem]:
    cols = _map_columns(header)
    items = []
    for i, row in enumerate(rows):
        get = lambda field: row[cols[field]] if field in cols and cols[field] < len(row) else None
        description = str(get("description") or "").strip()
        qty = _parse_int(get("qty"))
        unit_price = _parse_number(get("unit_price"))
        moq = _parse_int(get("moq"))
        lead_time = _parse_int(get("lead_time_days"))
        sku_code = (str(get("sku_code")).strip() or None) if get("sku_code") is not None else None
        items.append(LineItem(
            raw_description=description,
            qty=qty,
            unit_price=unit_price,
            moq=moq,
            lead_time_days=lead_time,
            brand=brands[i] if i < len(brands) else "UNKNOWN",
            vendor_gstin=vendor_gstin,
            source_file=source_file,
            source_format=source_format,
            sku_code=sku_code,
            extraction_confidence=_confidence(description, qty, unit_price),
            attrs=extract_attributes(description) if description else AttributeSignature(),
        ))
    return items


def parse_pdf(path: Path, vendor_gstin: str, brands: list[str]) -> list[LineItem]:
    header, rows = None, []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            if header is None:
                header, *body = table
            else:
                body = table
            rows.extend(body)
    return _rows_to_line_items(header or [], rows, vendor_gstin, brands, path.name, "pdf")


def parse_csv(path: Path, vendor_gstin: str, brands: list[str]) -> list[LineItem]:
    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)
    return _rows_to_line_items(header, rows, vendor_gstin, brands, path.name, "csv")


def parse_xlsx(path: Path, vendor_gstin: str, brands: list[str]) -> list[LineItem]:
    ws = load_workbook(path).active
    header, rows = None, []
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        if header is None:
            if len(_map_columns(list(row))) >= 2:
                header = list(row)
            continue
        if len(cells) == 1 and isinstance(cells[0], str) and cells[0].strip().startswith("*"):
            continue  # footnote row
        if not cells:
            continue
        rows.append(list(row))
    return _rows_to_line_items(header or [], rows, vendor_gstin, brands, path.name, "xlsx")


def _extract_commercial_terms(text: str) -> dict:
    qty_m, price_m, moq_m, lead_m = _QTY_RE.search(text), _PRICE_RE.search(text), _MOQ_RE.search(text), _LEAD_RE.search(text)
    return {
        "qty": _parse_int(qty_m.group(1)) if qty_m else None,
        "unit_price": _parse_number(price_m.group(1)) if price_m else None,
        "moq": _parse_int(moq_m.group(1)) if moq_m else None,
        "lead_time_days": _parse_int(lead_m.group(1)) if lead_m else None,
    }


def _blocks_to_line_items(blocks, vendor_gstin, brands, source_file, source_format, description_label) -> list[LineItem]:
    items = []
    for i, block in enumerate(blocks):
        terms = _extract_commercial_terms(block)
        if description_label:
            m = re.search(rf"{re.escape(description_label)}\s*(.+)", block)
            description = m.group(1).strip() if m else block.strip()
        else:
            description = block.strip()
        sku_m = _SKU_CODE_RE.search(block)
        # Chat sender tags (e.g. "[Rajkot Glass Works]:") are channel metadata,
        # not product spec - strip them before attribute extraction so a
        # vendor's own name can't leak a false material/colour/form token.
        attrs_text = _CHAT_TAG_RE.sub("", description)
        items.append(LineItem(
            raw_description=description,
            qty=terms["qty"],
            unit_price=terms["unit_price"],
            moq=terms["moq"],
            lead_time_days=terms["lead_time_days"],
            brand=brands[i] if i < len(brands) else "UNKNOWN",
            vendor_gstin=vendor_gstin,
            source_file=source_file,
            source_format=source_format,
            sku_code=sku_m.group(0) if sku_m else None,
            extraction_confidence=_confidence(description, terms["qty"], terms["unit_price"]),
            attrs=extract_attributes(attrs_text),
        ))
    return items


def parse_email(path: Path, vendor_gstin: str, brands: list[str]) -> list[LineItem]:
    text = path.read_text(encoding="utf-8")
    blocks = [b for b in re.split(r"\n-{3,}\n", text) if "Item:" in b]
    if not blocks:
        blocks = [text]
    return _blocks_to_line_items(blocks, vendor_gstin, brands, path.name, "email", description_label="Item:")


def parse_whatsapp(path: Path, vendor_gstin: str, brands: list[str]) -> list[LineItem]:
    text = path.read_text(encoding="utf-8")
    blocks = [b.strip() for b in re.split(r"\n-{3,}\n", text) if b.strip()]
    return _blocks_to_line_items(blocks, vendor_gstin, brands, path.name, "whatsapp", description_label=None)


_PARSERS = {"pdf": parse_pdf, "csv": parse_csv, "xlsx": parse_xlsx, "email": parse_email, "whatsapp": parse_whatsapp}


def extract_all(quotes_dir: Path = QUOTES_DIR) -> list[LineItem]:
    manifest = json.loads((quotes_dir / "manifest.json").read_text(encoding="utf-8"))
    items = []
    for entry in manifest:
        parser = _PARSERS[entry["format"]]
        items.extend(parser(quotes_dir / entry["source_file"], entry["vendor_gstin"], entry["line_item_brands"]))
    return items
