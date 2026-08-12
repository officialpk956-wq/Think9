"""Generates synthetic LOOM demo data: seed registries + 12 vendor quote files.

Run once: `python generate_data.py`. normalize.py's extraction rules were
written before this file existed (see CLAUDE.md build order) - nothing here
was reverse-engineered from what normalize.py happens to parse well.

Column headers vary deliberately per file (different wording, different
order) to exercise extract.py's schema-agnostic column matching rather than
positional parsing - that's the realistic version of "12 vendors, no two
quote formats alike."
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ROOT = Path(__file__).parent
QUOTES = ROOT / "data" / "quotes"
SEED = ROOT / "data" / "seed"

RAJKOT = "24AAACR5055K1Z5"
GUJARAT = "24AACCG9862P1ZX"
SHREE = "27AABCS1429B1ZQ"
DECCAN = "29AAGCD2287M1ZP"
MUMBAI = "27AAECM4521N1ZK"
EVEREST = "27AADCE7734L1ZR"

VENDORS = [
    {"gstin": RAJKOT, "name": "Rajkot Glass Works", "supplies": ["glass bottles"]},
    {"gstin": SHREE, "name": "Shree Packaging Industries", "supplies": ["caps", "pumps", "jars"]},
    {"gstin": GUJARAT, "name": "Gujarat Glass & Containers", "supplies": ["glass bottles"]},
    {"gstin": MUMBAI, "name": "Mumbai Flexipack", "supplies": ["pouches", "labels"]},
    {"gstin": EVEREST, "name": "Everest Cartons", "supplies": ["cartons"]},
    {"gstin": DECCAN, "name": "Deccan Closures", "supplies": ["caps", "pumps"]},
]

BRANDS = ["SuperYou", "Goodbug", "Panchamrit", "Neude", "Beauty by Bie"]

CANONICAL_SKUS = [
    {"id": "GLS-AMB-050-20", "description": "50ml amber glass bottle, 20mm neck", "moq": 2500,
     "attrs": {"material": "glass", "colour": "amber", "form": "bottle", "volume_ml": 50, "neck_mm": 20, "closure_included": False}},
    {"id": "GLS-AMB-100-24", "description": "100ml amber glass bottle, 24mm neck", "moq": 1500,
     "attrs": {"material": "glass", "colour": "amber", "form": "bottle", "volume_ml": 100, "neck_mm": 24}},
    {"id": "GLS-AMB-048-20", "description": "48ml amber glass bottle, 20mm neck", "moq": 3000,
     "attrs": {"material": "glass", "colour": "amber", "form": "bottle", "volume_ml": 48, "neck_mm": 20}},
    {"id": "CAP-PP-20-WHT", "description": "20mm white PP screw cap", "moq": 2000,
     "attrs": {"material": "PP", "colour": "white", "form": "cap", "neck_mm": 20}},
    {"id": "PMP-LOT-24-WHT", "description": "24mm white lotion pump", "moq": 600,
     "attrs": {"material": "PP", "colour": "white", "form": "pump", "neck_mm": 24}},
    {"id": "JAR-PET-200-70", "description": "200ml PET jar, 70mm mouth", "moq": 1000,
     "attrs": {"material": "PET", "colour": "clear", "form": "jar", "volume_ml": 200, "neck_mm": 70}},
    {"id": "CTN-KRF-090090120", "description": "Kraft carton 90x90x120mm", "moq": 5000,
     "attrs": {"material": "kraft", "form": "carton", "dimensions_mm": "90x90x120"}},
    {"id": "POU-STD-100-ZIP", "description": "100g stand-up zip pouch", "moq": 5000,
     "attrs": {"material": "laminate", "form": "pouch", "weight_g": 100}},
]

SKU_CODE_LOOKUP = [
    {"vendor_gstin": RAJKOT, "sku_code": "GB-AMB-50-20N", "canonical_id": "GLS-AMB-050-20"},
    {"vendor_gstin": DECCAN, "sku_code": "DCN-CAP20-WHT", "canonical_id": "CAP-PP-20-WHT"},
    {"vendor_gstin": DECCAN, "sku_code": "DCN-PMP24-WHT", "canonical_id": "PMP-LOT-24-WHT"},
]

# Most brand-SKU edges are direct buys. The exceptions are the interesting
# ones: a CM-embedded edge means the brand's contract manufacturer buys the
# component, so that volume cannot be consolidated into a Think9 PO without
# the CM's agreement - it gets rate-benchmarked instead of bundled.
PROCUREMENT_MODES = [
    {"brand": "Goodbug", "canonical_id": "POU-STD-100-ZIP", "procurement_mode": "cm_embedded"},
    {"brand": "SuperYou", "canonical_id": "CTN-KRF-090090120", "procurement_mode": "cm_embedded"},
    {"brand": "Panchamrit", "canonical_id": "CAP-PP-20-WHT", "procurement_mode": "unknown"},
]

# 90 days of historical POs - the cold-start backfill. Without this the
# price-outlier and lead-time-drift detectors have no baseline to compare
# against. (canonical, vendor, brand, unit_price, qty, lead_time_days, observed_at)
#
# lead_time_days is not in the spec's price_history table definition, but
# detector 4 compares a quoted lead time against "that vendor's trailing
# mean" - which has to come from somewhere. This is that somewhere.
PRICE_HISTORY = [
    ("GLS-AMB-050-20", RAJKOT, "Neude", 22.20, 500, 24, "2026-05-18"),
    ("GLS-AMB-050-20", RAJKOT, "Panchamrit", 21.80, 600, 25, "2026-06-12"),
    ("GLS-AMB-050-20", RAJKOT, "Beauty by Bie", 22.40, 450, 25, "2026-07-05"),
    ("GLS-AMB-050-20", GUJARAT, "Neude", 21.00, 700, 20, "2026-05-27"),
    ("GLS-AMB-050-20", GUJARAT, "Beauty by Bie", 20.90, 800, 20, "2026-06-20"),
    ("GLS-AMB-050-20", GUJARAT, "Neude", 21.20, 600, 21, "2026-07-08"),
    ("GLS-AMB-048-20", RAJKOT, "SuperYou", 19.20, 1000, 22, "2026-05-22"),
    ("GLS-AMB-048-20", RAJKOT, "SuperYou", 19.10, 1200, 22, "2026-06-30"),
    ("GLS-AMB-100-24", GUJARAT, "Panchamrit", 27.50, 300, 20, "2026-05-25"),
    ("GLS-AMB-100-24", GUJARAT, "Panchamrit", 27.80, 350, 20, "2026-07-02"),
    ("CAP-PP-20-WHT", SHREE, "Neude", 2.15, 500, 15, "2026-05-20"),
    ("CAP-PP-20-WHT", SHREE, "Beauty by Bie", 2.12, 550, 15, "2026-06-18"),
    ("CAP-PP-20-WHT", SHREE, "Neude", 2.14, 500, 16, "2026-07-09"),
    ("CAP-PP-20-WHT", DECCAN, "Panchamrit", 1.98, 600, 14, "2026-05-29"),
    ("CAP-PP-20-WHT", DECCAN, "Goodbug", 1.95, 700, 14, "2026-06-25"),
    ("CAP-PP-20-WHT", DECCAN, "Goodbug", 2.00, 500, 15, "2026-07-11"),
    ("PMP-LOT-24-WHT", SHREE, "Beauty by Bie", 8.60, 400, 17, "2026-05-19"),
    ("PMP-LOT-24-WHT", SHREE, "Neude", 8.55, 350, 18, "2026-06-15"),
    ("PMP-LOT-24-WHT", SHREE, "Beauty by Bie", 8.70, 400, 18, "2026-07-07"),
    ("PMP-LOT-24-WHT", DECCAN, "Neude", 7.80, 300, 15, "2026-06-05"),
    ("PMP-LOT-24-WHT", DECCAN, "Neude", 7.75, 300, 15, "2026-07-03"),
    ("JAR-PET-200-70", SHREE, "Neude", 11.50, 600, 15, "2026-05-21"),
    ("JAR-PET-200-70", SHREE, "Neude", 11.80, 500, 15, "2026-06-16"),
    ("JAR-PET-200-70", SHREE, "Neude", 12.00, 600, 16, "2026-07-06"),
    ("CTN-KRF-090090120", EVEREST, "SuperYou", 6.40, 2000, 12, "2026-05-16"),
    ("CTN-KRF-090090120", EVEREST, "Goodbug", 6.55, 1500, 12, "2026-06-11"),
    ("CTN-KRF-090090120", EVEREST, "Neude", 6.30, 1800, 13, "2026-07-04"),
    ("CTN-KRF-090090120", EVEREST, "Panchamrit", 6.45, 1000, 12, "2026-07-10"),
    ("POU-STD-100-ZIP", MUMBAI, "SuperYou", 3.10, 3000, 18, "2026-05-24"),
    ("POU-STD-100-ZIP", MUMBAI, "Goodbug", 3.15, 2500, 18, "2026-06-22"),
]

MANIFEST = []


def _write_pdf(filename, vendor_name, quote_date, valid_until, header, rows):
    doc = SimpleDocTemplate(str(QUOTES / filename), pagesize=A4)
    styles = getSampleStyleSheet()
    table = Table([header] + [[str(c) for c in row] for row in rows])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    doc.build([
        Paragraph(vendor_name, styles["Title"]),
        Paragraph(f"Quotation &nbsp;|&nbsp; Date: {quote_date} &nbsp;|&nbsp; Valid Until: {valid_until}", styles["Normal"]),
        Spacer(1, 12),
        table,
    ])


def _write_csv(filename, header, rows):
    with open(QUOTES / filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def _write_text(filename, body):
    (QUOTES / filename).write_text(body.strip() + "\n", encoding="utf-8")


def _write_xlsx(filename):
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:E1")
    ws["A1"] = "Neude - Vendor Comparison Sheet"
    ws.append(["Item", "Qty", "Rate", "MOQ", "Lead Time"])
    ws.append(["Airless Pump Bottle, 30ml, White", 250, 45.00, 1000, 20])
    ws.merge_cells("A4:E4")
    ws["A4"] = "*Rates subject to confirmation of final artwork; excludes GST."
    wb.save(QUOTES / filename)


def _manifest(source_file, vendor_gstin, issued_at, valid_until, fmt, brands):
    MANIFEST.append({
        "source_file": source_file,
        "vendor_gstin": vendor_gstin,
        "issued_at": issued_at,
        "valid_until": valid_until,
        "format": fmt,
        "line_item_brands": brands,
    })


def build():
    QUOTES.mkdir(parents=True, exist_ok=True)
    SEED.mkdir(parents=True, exist_ok=True)

    # 1. rajkot_glass_q3.pdf - hero item #1 + the 48ml near-miss control
    _write_pdf(
        "rajkot_glass_q3.pdf", "Rajkot Glass Works", "20 Jul 2026", "14 Aug 2026",
        ["Description", "Order Qty", "Rate (Rs)", "MOQ", "Lead Time (days)"],
        [
            # Neude's 800-unit requirement, also quoted by Gujarat in
            # email_gujarat_glass.txt - competing offers for one requirement,
            # not additive demand.
            ["50ml Amber Glass Bottle, 20mm neck", 800, "22.00", 2500, 25],
            ["48ml Amber Glass Bottle, 20mm neck", 1000, "19.00", 3000, 22],
        ],
    )
    _manifest("rajkot_glass_q3.pdf", RAJKOT, "2026-07-20", "2026-08-14", "pdf", ["Neude", "SuperYou"])

    # 2. gujarat_glass_quote.pdf - hero item #2 + single-brand 100ml
    _write_pdf(
        "gujarat_glass_quote.pdf", "Gujarat Glass & Containers", "28 Jul 2026", "20 Aug 2026",
        ["Item Description", "Qty Requested", "Unit Price INR", "Minimum Order Qty", "Lead Time"],
        [
            ["Amber Boston Round 50 ML (20/400)", 600, "21.50", 2500, 20],
            ["Amber Glass Bottle 100ml, 24mm neck", 300, "28.00", 1500, 20],
        ],
    )
    _manifest("gujarat_glass_quote.pdf", GUJARAT, "2026-07-28", "2026-08-20", "pdf", ["Beauty by Bie", "Panchamrit"])

    # 3. shree_packaging_rate_list.pdf - cap, pump (baseline lead time), jar
    _write_pdf(
        "shree_packaging_rate_list.pdf", "Shree Packaging Industries", "18 Jul 2026", "13 Aug 2026",
        ["Product", "Quantity", "Price/Unit (Rs)", "MOQ (units)", "Delivery (days)"],
        [
            ["20mm White PP Screw Cap", 500, "2.10", 2000, 15],
            ["24mm White Lotion Pump", 400, "8.50", 600, 18],
            ["200ml PET Jar, 70mm mouth", 600, "14.00", 1000, 15],
        ],
    )
    _manifest("shree_packaging_rate_list.pdf", SHREE, "2026-07-18", "2026-08-13", "pdf", ["Neude", "Beauty by Bie", "Neude"])

    # 4. everest_cartons.csv - 5-brand carton concentration risk
    _write_csv(
        "everest_cartons.csv", ["item", "qty", "rate", "moq"],
        [
            ["Kraft Carton 90x90x120mm", 2000, "6.50", 5000],
            ["Corrugated Carton Box 90x90x120", 1500, "6.80", 5000],
            ["Kraft Box 90 x 90 x 120 mm", 1000, "7.10", 5000],
            ["Carton 90x90x120", 1800, "6.60", 5000],
            ["Kraft Carton, 90x90x120mm, plain", 1200, "6.90", 5000],
        ],
    )
    _manifest("everest_cartons.csv", EVEREST, "2026-07-25", "2026-08-19", "csv",
              ["SuperYou", "Goodbug", "Panchamrit", "Neude", "Beauty by Bie"])

    # 5. mumbai_flexipack.csv - incompatible schema; pouch bundle + a novel label SKU
    _write_csv(
        "mumbai_flexipack.csv", ["Description", "Order Qty", "Price/Unit", "Min Order"],
        [
            ["100g Stand-up Pouch with Zip, Matte Laminate", 3000, "3.20", 5000],
            ["Zip Pouch 100gm Laminate Standup", 2500, "3.35", 5000],
            ["Printed BOPP Label Roll 50x30mm", 10000, "0.45", 20000],
        ],
    )
    _manifest("mumbai_flexipack.csv", MUMBAI, "2026-07-22", "2026-08-16", "csv", ["SuperYou", "Goodbug", "SuperYou"])

    # 6. deccan_closures.csv - third schema, vendor sku_code column enables Stage 1
    _write_csv(
        "deccan_closures.csv", ["sku_code", "product_desc", "units", "inr_per_unit"],
        [
            ["DCN-CAP20-WHT", "White PP Cap 20mm", 600, "1.95"],
            ["DCN-CAP20-WHT", "White PP Cap 20mm", 500, "2.00"],
            ["DCN-PMP24-WHT", "Lotion Pump 24mm White", 300, "7.75"],
        ],
    )
    _manifest("deccan_closures.csv", DECCAN, "2026-07-25", "2026-08-18", "csv", ["Panchamrit", "Goodbug", "Neude"])

    # 7. email_gujarat_glass.txt - hero item #4, prose pricing
    _write_text("email_gujarat_glass.txt", """
From: sales@gujaratglass.example
To: procurement@neude.example
Date: 22 Jul 2026
Subject: RE: Amber bottle enquiry

Hi team,

Following up on your enquiry - we can offer at Rs 20.75 per piece for 800 pcs.

Item: Glass bottle - amber - 50cc - neck 20mm - w/o cap
MOQ for this item is 2500 units. Lead time is 20 days.

Quote valid until 12 Aug 2026.

Regards,
Gujarat Glass & Containers
""")
    _manifest("email_gujarat_glass.txt", GUJARAT, "2026-07-22", "2026-08-12", "email", ["Neude"])

    # 8. email_shree_revised.txt - two items revised mid-thread; pump lead time drifts 18 -> 22 days
    _write_text("email_shree_revised.txt", """
From: sales@shreepackaging.example
To: procurement@beautybybie.example
Date: 02 Aug 2026
Subject: RE: Revised rates

Hi,

Following our call, please find revised rates below.

---

Item: 20mm White PP Screw Cap
Revised rate: Rs 2.05 per piece for 550 pcs. MOQ remains 2000 units. Lead time 16 days.

---

Item: 24mm White Lotion Pump
Revised rate: Rs 8.35 per piece for 400 pcs. MOQ 600 units. Lead time now 22 days.

---

Regards,
Shree Packaging Industries
""")
    _manifest("email_shree_revised.txt", SHREE, "2026-08-02", "2026-08-22", "email", ["Beauty by Bie", "Beauty by Bie"])

    # 9. whatsapp_rajkot.txt - hero item #3 via vendor sku_code, + the spec's own
    # textbook ambiguous case ("bottle - 50 - amber", no neck spec, no unit on the 50)
    _write_text("whatsapp_rajkot.txt", """
[Panchamrit Procurement]: Hi, need pricing for the amber bottles again
[Rajkot Glass Works]: Sure - GB-AMB-50-20N, same as last time, 450 pcs?
[Panchamrit Procurement]: Yes 450 pcs works, what's the rate
[Rajkot Glass Works]: Rs 22.50/pc, MOQ 2500 as before, lead time 25 days

---

[Panchamrit Procurement]: Also send rate for bottle - 50 - amber, need 300 pcs
[Rajkot Glass Works]: Rs 24/pc, MOQ 2500, lead time 25 days
""")
    _manifest("whatsapp_rajkot.txt", RAJKOT, "2026-08-01", "2026-08-26", "whatsapp", ["Panchamrit", "Panchamrit"])

    # 10. whatsapp_deccan.txt - clean cap reorder + a genuinely low-confidence message (no qty/price at all)
    _write_text("whatsapp_deccan.txt", """
[Goodbug Procurement]: Hi, need the usual white 20mm PP cap again, 700 pcs this time
[Deccan Closures]: Sure, same rate area - Rs 1.92/pc, MOQ 5000 as always, lead time 14 days

---

[SuperYou Procurement]: pump wala jo pehle bheja tha, rate thoda kam kar do
[Deccan Closures]: dekh ke batata hun
""")
    _manifest("whatsapp_deccan.txt", DECCAN, "2026-07-30", "2026-08-24", "whatsapp", ["Goodbug", "SuperYou"])

    # 11. whatsapp_goodbug_hindi.txt - hero item #5, Devanagari description
    _write_text("whatsapp_goodbug_hindi.txt", """
[Goodbug Procurement]: एम्बर ग्लास बॉटल 50ml 20mm - 700 pcs ka rate chahiye
[Rajkot Glass Works]: Rs 22.80/pc, MOQ 2500, lead time 25 days
""")
    _manifest("whatsapp_goodbug_hindi.txt", RAJKOT, "2026-08-03", "2026-08-28", "whatsapp", ["Goodbug"])

    # 12. neude_vendor_compare.xlsx - merged title + footnote rows; a genuinely novel SKU
    _write_xlsx("neude_vendor_compare.xlsx")
    _manifest("neude_vendor_compare.xlsx", SHREE, "2026-08-05", "2026-08-30", "xlsx", ["Neude"])

    (QUOTES / "manifest.json").write_text(json.dumps(MANIFEST, indent=2), encoding="utf-8")
    (SEED / "canonical_skus.json").write_text(json.dumps(CANONICAL_SKUS, indent=2), encoding="utf-8")
    (SEED / "brands.json").write_text(json.dumps(BRANDS, indent=2), encoding="utf-8")
    (SEED / "vendors.json").write_text(json.dumps(VENDORS, indent=2), encoding="utf-8")
    (SEED / "sku_code_lookup.json").write_text(json.dumps(SKU_CODE_LOOKUP, indent=2), encoding="utf-8")
    (SEED / "procurement_modes.json").write_text(json.dumps(PROCUREMENT_MODES, indent=2), encoding="utf-8")
    history = [
        {"canonical_id": c, "vendor_gstin": v, "brand": b, "unit_price": p,
         "qty": q, "lead_time_days": lt, "observed_at": obs}
        for c, v, b, p, q, lt, obs in PRICE_HISTORY
    ]
    (SEED / "price_history.json").write_text(json.dumps(history, indent=2), encoding="utf-8")

    print(f"Wrote {len(MANIFEST)} quote files to {QUOTES}")
    print(f"Wrote seed registries to {SEED}")


if __name__ == "__main__":
    build()
